from __future__ import annotations

import csv
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))

from backend.src.downloaders.ofz_auctions_downloader import RAW_FILES_DIR
from backend.src.downloaders.ofz_auctions_downloader import RAW_INDEX_FILE
from backend.src.downloaders.ofz_auctions_downloader import parse_document_index

OUTPUT_FILE = PROJECT_ROOT / "data/processed/ofz_auctions.csv"

OUTPUT_COLUMNS = [
    "auction_date",
    "published_date",
    "document_title",
    "auction_format",
    "issue",
    "security_type",
    "maturity_date",
    "days_to_maturity",
    "offered_amount",
    "demand_amount",
    "placed_amount",
    "proceeds_amount",
    "cutoff_price",
    "weighted_average_price",
    "cutoff_yield",
    "weighted_average_yield",
    "official_coefficient",
    "cover_ratio",
    "placement_ratio",
    "source_url",
    "source_file",
]

HEADER_TO_FIELD = {
    "дата": "auction_date",
    "дата аукциона": "auction_date",
    "формат": "auction_format",
    "код выпуска": "issue",
    "тип бумаги": "security_type",
    "дата погашения": "maturity_date",
    "дней до погашения": "days_to_maturity",
    "объем предложения": "offered_amount",
    "цена отсечения": "cutoff_price",
    "цена средневзвешенная": "weighted_average_price",
    "доходность по цене отсечения": "cutoff_yield",
    "доходность по средневзвешенной цене": "weighted_average_yield",
    "совокупный объем спроса по номиналу": "demand_amount",
    "объем размещения по номиналу": "placed_amount",
    "объем выручки": "proceeds_amount",
    "коэффициент активности": "official_coefficient",
    "коэффициент размещения на аукционе": "official_coefficient",
    "коэффициент удовлетворения спроса на аукционе": "official_coefficient",
}


def _normalize_space(value: str) -> str:
    """Нормализует пробелы в строке"""
    return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()


def _normalize_header(value: object) -> str:
    """Нормализует заголовок Excel-колонки"""
    text = _normalize_space(str(value or "")).lower()
    text = re.sub(r"\*+", "", text)
    text = text.replace("средневзве- шенной", "средневзвешенной")
    text = re.sub(r"[^а-яa-z0-9\s]", " ", text, flags=re.IGNORECASE)
    return _normalize_space(text)


def _format_date(value: object) -> str:
    """Преобразует дату в формат DD-MM-YYYY"""
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if value is None:
        return ""

    text = _normalize_space(str(value))
    for date_format in ("%d.%m.%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, date_format).strftime("%d-%m-%Y")
        except ValueError:
            continue
    return ""


def _to_float(value: object) -> float | None:
    """Преобразует значение Excel в число с плавающей точкой"""
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None

    text = _normalize_space(str(value))
    text = text.replace(" ", "").replace(",", ".")
    if text in {"", "-", "—"} or text.startswith("-*"):
        return None

    try:
        return float(text)
    except ValueError:
        return None


def _to_int(value: object) -> int | None:
    """Преобразует значение Excel в целое число"""
    number = _to_float(value)
    if number is None:
        return None
    return int(number)


def _safe_divide(numerator: float | None, denominator: float | None) -> float | None:
    """Делит числа с защитой от пустого или нулевого знаменателя"""
    if numerator is None or denominator in {None, 0}:
        return None
    return numerator / denominator


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    """Находит строку заголовков в Excel-файле Минфина"""
    for index, row in enumerate(rows):
        headers = {_normalize_header(value) for value in row}
        if "объем предложения" in headers and "код выпуска" in headers:
            return index
    raise ValueError("Не найдена строка заголовков в Excel-файле Минфина")


def _build_column_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    """Сопоставляет заголовки Excel с внутренними именами колонок"""
    columns: dict[str, int] = {}
    for index, value in enumerate(header_row):
        field_name = HEADER_TO_FIELD.get(_normalize_header(value))
        if field_name is not None and field_name not in columns:
            columns[field_name] = index
    return columns


def _get(row: tuple[Any, ...], columns: dict[str, int], field_name: str) -> object:
    """Возвращает значение ячейки по имени поля"""
    index = columns.get(field_name)
    if index is None or index >= len(row):
        return None
    return row[index]


def _parse_workbook(
    path: Path,
    document_title: str,
    source_url: str,
    published_date: str,
) -> list[dict[str, object]]:
    """Парсит один XLSX-файл Минфина с результатами аукционов"""
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    header_index = _find_header_row(rows)
    columns = _build_column_map(rows[header_index])

    parsed_rows: list[dict[str, object]] = []
    for row in rows[header_index + 1 :]:
        auction_date = _format_date(_get(row, columns, "auction_date"))
        issue = _normalize_space(str(_get(row, columns, "issue") or ""))
        if not auction_date or not issue or issue.lower() == "итого":
            continue

        offered_amount = _to_float(_get(row, columns, "offered_amount"))
        demand_amount = _to_float(_get(row, columns, "demand_amount"))
        placed_amount = _to_float(_get(row, columns, "placed_amount"))

        parsed_rows.append(
            {
                "auction_date": auction_date,
                "published_date": _format_date(published_date),
                "document_title": document_title,
                "auction_format": _normalize_space(
                    str(_get(row, columns, "auction_format") or "")
                ),
                "issue": issue,
                "security_type": _normalize_space(
                    str(_get(row, columns, "security_type") or "")
                ),
                "maturity_date": _format_date(_get(row, columns, "maturity_date")),
                "days_to_maturity": _to_int(_get(row, columns, "days_to_maturity")),
                "offered_amount": offered_amount,
                "demand_amount": demand_amount,
                "placed_amount": placed_amount,
                "proceeds_amount": _to_float(_get(row, columns, "proceeds_amount")),
                "cutoff_price": _to_float(_get(row, columns, "cutoff_price")),
                "weighted_average_price": _to_float(
                    _get(row, columns, "weighted_average_price")
                ),
                "cutoff_yield": _to_float(_get(row, columns, "cutoff_yield")),
                "weighted_average_yield": _to_float(
                    _get(row, columns, "weighted_average_yield")
                ),
                "official_coefficient": _to_float(
                    _get(row, columns, "official_coefficient")
                ),
                "cover_ratio": _safe_divide(demand_amount, offered_amount),
                "placement_ratio": _safe_divide(placed_amount, offered_amount),
                "source_url": source_url,
                "source_file": path.name,
            }
        )

    return parsed_rows


def _date_sort_key(value: object) -> datetime:
    """Готовит дату DD-MM-YYYY для сортировки"""
    return datetime.strptime(str(value), "%d-%m-%Y")


def parse_ofz_auctions(
    index_path: Path = RAW_INDEX_FILE,
    files_dir: Path = RAW_FILES_DIR,
) -> list[dict[str, object]]:
    """Парсит результаты аукционов ОФЗ из XLSX-файлов Минфина"""
    documents = parse_document_index(index_path)

    parsed_rows: list[dict[str, object]] = []
    for document in documents:
        file_path = files_dir / document.file_name
        if not file_path.exists():
            continue

        parsed_rows.extend(
            _parse_workbook(
                file_path,
                document.title,
                document.document_url,
                document.published_date,
            )
        )

    return sorted(
        parsed_rows,
        key=lambda row: (
            _date_sort_key(row["auction_date"]),
            str(row["issue"]),
            str(row["auction_format"]),
        ),
    )


def save_csv(rows: list[dict[str, object]], output_path: Path = OUTPUT_FILE) -> None:
    """Сохраняет подготовленные строки в CSV-файл"""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    """Запускает парсер и сохраняет результат в CSV"""
    rows = parse_ofz_auctions()
    save_csv(rows)
    print(f"Сохранено строк: {len(rows)}")
    print(f"Файл: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
